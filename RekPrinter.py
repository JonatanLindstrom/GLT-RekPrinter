import os
import re
import datetime
# Import Workbook
from openpyxl import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
# Import GUI
from tkinter import *
from tkinter import filedialog


def getPath():
    root.filename = filedialog.askopenfilename(title = 'Select file', filetypes = (('Excel files','*.xlsx'),('All files','*.*')))
    return root.filename


def setPath(path):
    now = datetime.datetime.fromtimestamp(os.path.getctime(path)).strftime('%Y%m%d')
    path = path.split('/')
    output = ''
    for i in range(len(path) - 1):
        output += path[i] + '/'
    output += ('Kvällsbeställningar ' + now[2:] + '.xlsx')

    i = 0
    while os.path.isfile(output):
        if i == 0:
            output = output[:-5] + ' (1).xlsx'
        else:
            parenthesis = output.find('(')
            output = output[:parenthesis+1] + str(i) + ').xlsx'
        i += 1

    return output


def copyRow(row, startCol, endCol, sheet):
    rowSelected = []
    for j in range(startCol,endCol+1):
        rowSelected.append(sheet.cell(row = row, column = j).value)

    return rowSelected


def pasteRow(row, startCol, endCol, sheetReceiving, copiedData):
    countCol = 0
    for j in range(startCol,endCol+1):
        sheetReceiving.cell(row = row, column = j).value = copiedData[countCol]
        countCol += 1

        sheetReceiving.cell(row = row, column = j).border = Border(left=Side(border_style='thin', color='00000000'),
                                                                    right=Side(border_style='thin', color='00000000'),
                                                                    top=Side(border_style='thin', color='00000000'),
                                                                    bottom=Side(border_style='thin', color='00000000'))

    sheetReceiving.column_dimensions['A'].width = 6
    sheetReceiving.column_dimensions['B'].width = 6
    sheetReceiving.column_dimensions['C'].width = 15
    sheetReceiving.column_dimensions['D'].width = 39
    sheetReceiving.column_dimensions['E'].width = 9
    sheetReceiving.column_dimensions['F'].width = 6


def splitReq(wb):
    orgWS = wb.active
    orgWS.delete_cols(5)
    orgWS.delete_cols(2)

    for placeRow in placeMap:
        wb.create_sheet(placeRow[1])

    newWSi = 0
    orgWSi = 0
    previousRow = ['Radnr', 'Företagskod', 'Företag', 'Benämning', 'Återstår antal', 'Enhet']
    for row in orgWS.rows:
        rowlist = list()
        for cell in row:
            rowlist.append(str(cell.value))
        
        orgWSi += 1
        if rowlist[2] != previousRow[2]:
            newWSi = 1
        else:
            newWSi += 1    

        if rowlist[2] != 'Företag':
            activeWS = wb[rowlist[2].replace(':', '')]
            activeRow = copyRow(orgWSi, 1, len(rowlist), orgWS)
            pasteRow(newWSi, 1, len(activeRow), activeWS, activeRow)

        previousRow = rowlist
    
    wb.remove(orgWS)


def checkReq(wb):
    for sheetName in wb.sheetnames:
        if sheetName == 'Saknade rekar':
            continue
        sheet = wb[sheetName]
        if sheet['A1'].value == None:
            wb.remove(sheet)

    reqs = wb.sheetnames
    missing = list()
    for row in placeMap:
        place = row[1]
        if place not in reqs:
            if re.match('.* [0-9]an', place):
                row[1] = place[:-2] + ':' + place[-2:]
                missing.append(row)
            else:
                missing.append(row)
    
    wb.create_sheet('Saknade rekar', 0)
    activeWS = wb['Saknade rekar']
    activeWS['C1'] = 'Saknade rekar'
    activeWS['C1'].alignment = Alignment(horizontal='center')
    activeWS.merge_cells('C1:E1')

    pasteRow(3, 3, 5, activeWS, ['Kostnadsställe', 'Enhet', 'Telefon'])
    i = 5
    for row in missing:
        pasteRow(i, 3, 5, activeWS, row)
        i += 1


def formatFile(path):
    wb = load_workbook(filename=path)

    splitReq(wb)
    
    checkReq(wb)

    output = setPath(path)
    wb.save(output)
    os.startfile(output)


def main():
    path = getPath()
    formatFile(path)


root = Tk()
root.withdraw()

placeMap = [['564', 'Ham 1an', 9703],
            ['567', 'Glass & Pop 1an', 9704],
            ['574', 'Tivolisnacks', 9702],
            ['551', 'Slushbaren', 9705],
            ['571', 'Honeycomb', 0000],
            ['581', 'Korvvagn 1', 9707],
            ['575', 'Remvagn 1', 9706],
            ['584', 'Coca cola store', 9807],
            ['563', 'Matvraket', 9718],
            ['576', 'Kebaben', 9714],
            ['562', 'Mexican Corner', 9711],
            ['572', 'Pizzan', 9710],
            ['565', 'Korv 2an', 9716],
            ['561', 'Coffeebar', 9712],
            ['554', 'Pop 2an', 9715],
            ['568', 'Glass 2an', 9713],
            ['570', 'Grädderiet', 9719],
            ['578', 'Remvagn 2', 9720],
            ['560', 'Godisfabriken', 9717],
            ['566', 'Hamburger 3an', 9737],
            ['569', 'Gyros', 9736],
            ['558', 'Langos', 9734],
            ['559', 'Fish & Chips', 9734],
            ['573', 'Poké Bowl', 9734],
            ['557', 'Coffee and Donuts', 9730],
            ['556', 'Glasskammaren', 9731],
            ['553', 'Pop 3an', 9732],
            ['582', 'Milkshakebaren', 9810],
            ['577', 'Kvastenkiosken', 9735],
            ['555', 'Boardwalk Café', 9733],
            ['580', 'Popcorn & Cotton Candy', 9733],
            ['579', 'Remvagn 3', 9739],
            ['583', 'Korvvagn 3', 9740],
            ['552', 'Ben & Jerry\'s', 9738],
            ['612', '1883-butiken', 9801],
            ['613', 'Tivolibutiken', 9803],
            ['623', 'Fotobutik Lustiga huset', 9808],
            ['626', 'Fotobutik Twister', 9809]]


if __name__ == '__main__':
    main()

